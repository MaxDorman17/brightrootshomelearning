import re
import io
import json
import asyncio
import httpx
import concurrent.futures
from datetime import date
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from pydantic import BaseModel
from sqlalchemy import or_, and_
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from auth import get_current_user, require_parent
from database import get_db, SessionLocal
from models import User, OakQuizResult, PlannerEntry, PlannerCompletion, Lesson
from config import settings

router = APIRouter(prefix="/api/oak", tags=["oak"])

OAK_HEADERS = {
    "Authorization": f"Bearer {settings.OAK_API_KEY}",
    "Accept": "application/json",
}

YEAR_TO_KS = {
    "Year 1": "key stage 1", "Year 2": "key stage 1",
    "Year 3": "key stage 2", "Year 4": "key stage 2",
    "Year 5": "key stage 2", "Year 6": "key stage 2",
    "Year 7": "key stage 3", "Year 8": "key stage 3", "Year 9": "key stage 3",
    "Year 10": "key stage 4", "Year 11": "key stage 4",
}

# Topic keywords per subject — each triggers a separate parallel search
SUBJECT_KEYWORDS = {
    "maths": ["number", "algebra", "geometry", "ratio", "statistics", "fractions", "decimals", "probability", "angles", "sequences"],
    "english": ["reading", "writing", "poetry", "grammar", "fiction", "non-fiction", "shakespeare", "narrative", "vocabulary", "punctuation"],
    "science": ["biology", "chemistry", "physics", "cells", "forces", "energy", "atoms", "ecology", "electricity", "waves"],
    "history": ["medieval", "empire", "revolution", "monarchy", "war", "ancient", "migration", "power", "society", "crime"],
    "geography": ["rivers", "climate", "urbanisation", "development", "coasts", "ecosystems", "population", "resources", "weather", "globalisation"],
    "art and design": ["drawing", "painting", "sculpture", "printmaking", "textiles", "colour", "composition", "design", "portrait", "abstract"],
    "design and technology": ["engineering", "product", "food", "textiles", "resistant materials", "structures", "mechanisms", "electronics", "cooking", "materials"],
    "computing": ["programming", "algorithms", "data", "networks", "cybersecurity", "python", "html", "binary", "database", "software"],
    "religious education": ["christianity", "islam", "buddhism", "hinduism", "ethics", "beliefs", "worship", "sacred", "morality", "philosophy"],
    "physical education": ["fitness", "athletics", "team", "skills", "health", "swimming", "gymnastics", "dance", "tactics", "movement"],
    "music": ["rhythm", "melody", "harmony", "composition", "notation", "instruments", "performance", "listening", "blues", "classical"],
    "french": ["vocabulary", "grammar", "reading", "listening", "speaking", "writing", "verbs", "nouns", "phrases", "culture"],
    "spanish": ["vocabulary", "grammar", "reading", "listening", "speaking", "writing", "verbs", "nouns", "phrases", "culture"],
    "german": ["vocabulary", "grammar", "reading", "listening", "speaking", "writing", "verbs", "nouns", "phrases", "culture"],
}


def _fetch_one(query: str) -> list:
    url = f"{settings.OAK_BASE_URL}/search/lessons"
    try:
        with httpx.Client(timeout=10) as client:
            r = client.get(url, headers=OAK_HEADERS, params={"q": query})
            r.raise_for_status()
            return r.json().get("hits", {}).get("hits", [])
    except Exception:
        return []


def build_oak_url(src: dict) -> str:
    subject_slug = src.get("subject_slug", "")
    phase = src.get("phase", "primary")
    ks_slug = src.get("key_stage_slug", "")
    ks_abbrev = ks_slug.replace("key-stage-", "ks")
    programme_slug = f"{subject_slug}-{phase}-{ks_abbrev}"
    unit_slug = src.get("topic_slug", "")
    lesson_slug = src.get("slug", "")
    return f"https://www.thenational.academy/pupils/programmes/{programme_slug}/units/{unit_slug}/lessons/{lesson_slug}"


@router.get("/search")
def search_lessons(
    q: str = Query(default="", min_length=0),
    subject: str = Query(default=""),
    year: str = Query(default=""),
    _: User = Depends(get_current_user),
):
    ks = YEAR_TO_KS.get(year, "")
    subject_lower = subject.lower()
    keywords = SUBJECT_KEYWORDS.get(subject_lower, [""])

    # Build one query per keyword — run all in parallel
    queries = [
        f"{subject} {year} {ks} {kw} {q}".strip()
        for kw in keywords
    ]

    seen_slugs: dict = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(_fetch_one, qry) for qry in queries]
        for future in concurrent.futures.as_completed(futures):
            for hit in future.result():
                src = hit.get("_source", {})
                slug = src.get("slug")
                if slug and slug not in seen_slugs:
                    seen_slugs[slug] = src

    results = []
    for src in seen_slugs.values():
        if year and src.get("year_title", "") != year:
            continue
        if subject and src.get("subject_title", "").lower() != subject_lower:
            continue

        results.append({
            "slug": src.get("slug"),
            "title": src.get("title", ""),
            "subject": src.get("subject_title", ""),
            "keyStage": src.get("key_stage_title", ""),
            "yearTitle": src.get("year_title", ""),
            "unitTitle": src.get("topic_title", ""),
            "description": src.get("lesson_description", ""),
            "oakUrl": build_oak_url(src),
        })

    results.sort(key=lambda r: (r["unitTitle"], r["title"]))
    return results


class ImportUnitRequest(BaseModel):
    unit_url: str


@router.post("/import-unit")
async def import_unit(
    body: ImportUnitRequest,
    current_user: User = Depends(require_parent),
):
    match = re.search(r"/programmes/([^/?#]+)/units/([^/?#]+)", body.unit_url)
    if not match:
        raise HTTPException(
            status_code=400,
            detail="Could not parse unit URL — paste a thenational.academy unit link",
        )
    programme_slug = match.group(1)
    unit_slug = match.group(2)

    # Fetch the public pupil unit page and extract __NEXT_DATA__
    lessons_url = (
        f"https://www.thenational.academy/pupils/programmes/"
        f"{programme_slug}/units/{unit_slug}/lessons"
    )
    async with httpx.AsyncClient(follow_redirects=True, timeout=20.0) as client:
        try:
            resp = await client.get(
                lessons_url,
                headers={"User-Agent": "Mozilla/5.0 (compatible; HomeschoolApp/1.0)"},
            )
        except httpx.RequestError as exc:
            raise HTTPException(status_code=502, detail=f"Could not reach Oak website: {exc}")

    if resp.status_code != 200:
        raise HTTPException(status_code=404, detail="Unit not found — check the URL")

    m = re.search(
        r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>',
        resp.text,
        re.DOTALL,
    )
    if not m:
        raise HTTPException(status_code=502, detail="Could not parse lesson data from Oak page")

    try:
        page_data = json.loads(m.group(1))
    except json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="Could not parse lesson data from Oak page")

    browse_data = page_data.get("props", {}).get("pageProps", {}).get("browseData", [])
    if not browse_data:
        raise HTTPException(status_code=404, detail="No lesson data found on this page")

    first = browse_data[0]
    lesson_list = first.get("supplementaryData", {}).get("staticLessonList", [])
    unit_title = first.get("unitData", {}).get("title", "")

    lessons = []
    for lesson in lesson_list:
        if lesson.get("_state") == "published":
            slug = lesson.get("slug", "")
            title = lesson.get("title", "")
            order = lesson.get("order", 0)
            if slug and title:
                pupil_url = (
                    f"https://www.thenational.academy/pupils/programmes/"
                    f"{programme_slug}/units/{unit_slug}/lessons/{slug}"
                )
                lessons.append({"title": title, "url": pupil_url, "order": order})

    lessons.sort(key=lambda x: x["order"])
    return {
        "lessons": lessons,
        "unit_slug": unit_slug,
        "unit_title": unit_title,
        "programme_slug": programme_slug,
    }


# ---------------------------------------------------------------------------
# Quiz results from Oak "share my results" links
# ---------------------------------------------------------------------------

OAK_SHARE_RE = re.compile(
    r"https?://(?:www\.)?thenational\.academy/pupils/lessons/[^/?#]+/results/[^/?#]+/share"
)

SHARE_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; HomeschoolApp/1.0)"}


def _extract_quiz_scores(html: str) -> dict | None:
    """Pull starter/exit quiz scores out of the share page's embedded JSON.

    The page contains: "sectionResults":{..."starter-quiz":{"grade":2,"numQuestions":6,...},
    "exit-quiz":{"grade":5,"numQuestions":6,...}}
    """
    anchor = html.find('"sectionResults"')
    if anchor == -1:
        return None
    region = html[anchor:]

    def section(name: str):
        i = region.find(f'"{name}"')
        if i == -1:
            return None, None
        chunk = region[i:i + 200]
        grade = re.search(r'"grade":(\d+)', chunk)
        total = re.search(r'"numQuestions":(\d+)', chunk)
        return (
            int(grade.group(1)) if grade else None,
            int(total.group(1)) if total else None,
        )

    starter_score, starter_total = section("starter-quiz")
    exit_score, exit_total = section("exit-quiz")
    if starter_total is None and exit_total is None:
        return None
    return {
        "starter_score": starter_score,
        "starter_total": starter_total,
        "exit_score": exit_score,
        "exit_total": exit_total,
    }


async def _fetch_share_scores(client: httpx.AsyncClient, url: str) -> dict | None:
    try:
        resp = await client.get(url, headers=SHARE_HEADERS)
        if resp.status_code != 200:
            return None
        return _extract_quiz_scores(resp.text)
    except httpx.RequestError:
        return None


def _upsert_result(db: Session, url: str, scores: dict) -> None:
    row = db.query(OakQuizResult).filter(OakQuizResult.url == url).first()
    if row:
        for k, v in scores.items():
            setattr(row, k, v)
    else:
        db.add(OakQuizResult(url=url, **scores))
    db.commit()


async def fetch_and_store_share_result(url: str) -> None:
    """Background task: fetch one share link and cache its quiz scores."""
    async with httpx.AsyncClient(follow_redirects=True, timeout=20.0) as client:
        scores = await _fetch_share_scores(client, url)
    if scores is None:
        return
    db = SessionLocal()
    try:
        _upsert_result(db, url, scores)
    finally:
        db.close()


@router.get("/quiz-results")
def get_quiz_results(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    rows = db.query(OakQuizResult).all()
    return [
        {
            "url": r.url,
            "starter_score": r.starter_score,
            "starter_total": r.starter_total,
            "exit_score": r.exit_score,
            "exit_total": r.exit_total,
        }
        for r in rows
    ]


@router.post("/quiz-results/refresh")
async def refresh_quiz_results(
    db: Session = Depends(get_db),
    _: User = Depends(require_parent),
):
    """Fetch quiz scores for any submitted Oak share links not yet cached."""
    urls: set = set()
    for (u,) in db.query(PlannerEntry.completed_work_url).filter(
        PlannerEntry.completed_work_url.is_not(None)
    ).all():
        if u and OAK_SHARE_RE.search(u):
            urls.add(OAK_SHARE_RE.search(u).group(0))
    for (u,) in db.query(PlannerCompletion.completed_work_url).filter(
        PlannerCompletion.completed_work_url.is_not(None)
    ).all():
        if u and OAK_SHARE_RE.search(u):
            urls.add(OAK_SHARE_RE.search(u).group(0))

    cached = {r.url for r in db.query(OakQuizResult).all()}
    missing = sorted(urls - cached)

    added = 0
    if missing:
        sem = asyncio.Semaphore(6)
        async with httpx.AsyncClient(follow_redirects=True, timeout=20.0) as client:
            async def fetch_one(u: str):
                async with sem:
                    return u, await _fetch_share_scores(client, u)

            results = await asyncio.gather(*(fetch_one(u) for u in missing))
        for u, scores in results:
            if scores is not None:
                _upsert_result(db, u, scores)
                added += 1

    return {"checked": len(missing), "added": added, "total_cached": len(cached) + added}


# ---------------------------------------------------------------------------
# Excel export of Oak quiz results — homeschool evidence record
# ---------------------------------------------------------------------------

EXPORT_HEADERS = [
    "Scheduled Date", "Completed Date", "Child", "Subject", "Lesson Title",
    "Assignment Type", "Starter Score", "Starter Total", "Starter %",
    "Exit Score", "Exit Total", "Exit %", "Oak Results URL", "Note",
]
EXPORT_COLUMN_WIDTHS = [14, 18, 14, 16, 34, 14, 12, 12, 10, 10, 10, 8, 48, 30]


@router.get("/export")
async def export_oak_results(
    child_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_parent),
):
    """Export this parent's children's Oak quiz results as a formatted .xlsx.

    Directly-assigned entries are attributed via PlannerEntry.assigned_to.
    Shared (assigned_to=NULL) entries are attributed by enumerating every
    matching PlannerCompletion row individually — one export row per child
    completion — so a second child's submission is never silently dropped
    the way a single 'best' pick would. All queries are scoped to this
    parent's own children (direct) or their own lessons (shared), so one
    family's export can never include another family's data.
    """
    children = db.query(User).filter(User.parent_id == current_user.id).all()
    child_ids = [c.id for c in children]
    child_names = {c.id: c.username for c in children}

    if child_id is not None and child_id not in child_ids:
        raise HTTPException(status_code=403, detail="Not your child")
    target_child_ids = [child_id] if child_id is not None else child_ids

    query = db.query(PlannerEntry).join(Lesson, PlannerEntry.lesson_id == Lesson.id).options(
        joinedload(PlannerEntry.lesson)
    ).filter(
        or_(
            PlannerEntry.assigned_to.in_(target_child_ids),
            and_(PlannerEntry.assigned_to.is_(None), Lesson.created_by == current_user.id),
        )
    )
    if start_date:
        query = query.filter(PlannerEntry.scheduled_date >= start_date)
    if end_date:
        query = query.filter(PlannerEntry.scheduled_date <= end_date)
    entries = query.order_by(PlannerEntry.scheduled_date).all()

    # Bulk-fetch every PlannerCompletion for the shared entries in this set —
    # never just the 'best' one — scoped to this parent's own children.
    shared_ids = [e.id for e in entries if e.assigned_to is None]
    comps_by_entry: dict = {}
    if shared_ids and target_child_ids:
        for comp in db.query(PlannerCompletion).filter(
            PlannerCompletion.entry_id.in_(shared_ids),
            PlannerCompletion.user_id.in_(target_child_ids),
            PlannerCompletion.completed_work_url.is_not(None),
        ).all():
            comps_by_entry.setdefault(comp.entry_id, []).append(comp)

    # Build one candidate row per (entry, child) submission that has an Oak share link.
    candidates = []
    for e in entries:
        if e.assigned_to is not None:
            url = e.completed_work_url
            match = OAK_SHARE_RE.search(url) if url else None
            if match:
                candidates.append({
                    "child_id": e.assigned_to,
                    "url": match.group(0),
                    "note": e.completed_note,
                    "completed_at": e.completed_at,
                    "scheduled_date": e.scheduled_date,
                    "lesson": e.lesson,
                    "assignment_type": "Direct",
                })
        else:
            for comp in comps_by_entry.get(e.id, []):
                url = comp.completed_work_url
                match = OAK_SHARE_RE.search(url) if url else None
                if match:
                    candidates.append({
                        "child_id": comp.user_id,
                        "url": match.group(0),
                        "note": comp.completed_note,
                        "completed_at": comp.completed_at,
                        "scheduled_date": e.scheduled_date,
                        "lesson": e.lesson,
                        "assignment_type": "Shared",
                    })

    # Reuse the existing share-page fetch/cache logic (same as quiz-results/refresh)
    # rather than writing a second scraper — scoped to just this export's URLs.
    canonical_urls = {c["url"] for c in candidates}
    cached = {
        r.url: r for r in db.query(OakQuizResult).filter(OakQuizResult.url.in_(canonical_urls)).all()
    } if canonical_urls else {}
    missing = sorted(canonical_urls - cached.keys())
    if missing:
        sem = asyncio.Semaphore(6)
        async with httpx.AsyncClient(follow_redirects=True, timeout=20.0) as client:
            async def fetch_one(u: str):
                async with sem:
                    return u, await _fetch_share_scores(client, u)

            results = await asyncio.gather(*(fetch_one(u) for u in missing))
        for u, scores in results:
            if scores is not None:
                _upsert_result(db, u, scores)
        cached = {
            r.url: r for r in db.query(OakQuizResult).filter(OakQuizResult.url.in_(canonical_urls)).all()
        }

    candidates.sort(key=lambda c: (c["scheduled_date"], child_names.get(c["child_id"], "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Oak Results"
    ws.append(EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for c in candidates:
        result = cached.get(c["url"])
        starter_score = result.starter_score if result else None
        starter_total = result.starter_total if result else None
        exit_score = result.exit_score if result else None
        exit_total = result.exit_total if result else None
        starter_pct = (starter_score / starter_total) if (starter_score is not None and starter_total) else None
        exit_pct = (exit_score / exit_total) if (exit_score is not None and exit_total) else None
        completed_at = c["completed_at"]
        if completed_at is not None and completed_at.tzinfo is not None:
            completed_at = completed_at.replace(tzinfo=None)  # Excel doesn't support tz-aware datetimes

        ws.append([
            c["scheduled_date"],
            completed_at,
            child_names.get(c["child_id"], "Unknown"),
            c["lesson"].subject,
            c["lesson"].title,
            c["assignment_type"],
            starter_score,
            starter_total,
            starter_pct,
            exit_score,
            exit_total,
            exit_pct,
            c["url"],
            c["note"] or "",
        ])

    last_row = ws.max_row
    if last_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = "yyyy-mm-dd"
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm"
        for col_idx in (9, 12):  # Starter %, Exit %
            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = "0%"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, width in enumerate(EXPORT_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bright-roots-oak-results-{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
